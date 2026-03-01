"""
Online Attendance Analysis Tool
Web-based application for PDF to Excel conversion
"""

import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
import sys
import io
import tempfile
import os
import plotly.graph_objects as go
import plotly.express as px

# Import the existing processing classes
sys.path.append('..')
from excel_consolidator import ExcelConsolidator
from pdf_extractor import AttendancePDFExtractor, RosterExtractor
from report_generator import TrainOperationsRosterReportGenerator

# Optional: try to import Night Shift report model from parent project
night_shift_model = None
try:
    # Common possible module names - adjust if your model file has different name
    from night_shift_report import generate_night_shift_report
    night_shift_model = generate_night_shift_report
except Exception:
    try:
        from nightshift import generate_night_shift_report
        night_shift_model = generate_night_shift_report
    except Exception:
        night_shift_model = None

# Page configuration
st.set_page_config(
    page_title="Attendance Analysis Tool",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS (same as dashboard)
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        padding: 1rem;
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        color: white;
        border-radius: 10px;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #1f77b4;
    }
    .stAlert {
        border-radius: 10px;
    }
    </style>
""", unsafe_allow_html=True)


def process_pdf_to_excel(uploaded_file, department='Stations', pdf_type='attendance'):
    """
    Process uploaded PDF file and return Excel data
    
    Args:
        uploaded_file: The uploaded PDF file
        department: 'Stations', 'OCC', or 'Train Operations'
        pdf_type: Type of PDF - 'attendance' for Stations/OCC/Roster, 'trip_chart' for Train Operations Trip Chart
    """
    try:
        # Save uploaded file to temporary location
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            tmp_path = tmp_file.name

        # Choose extractor based on department and pdf_type
        if pdf_type == 'trip_chart':
            # Trip Chart: use AttendancePDFExtractor (with Shift/Shift_Timings)
            extractor = AttendancePDFExtractor(tmp_path)
            extractor.extract_tables()
            df = extractor.create_dataframe()
            sheet_name = 'Trip Chart Data'
        else:
            # Stations, OCC, Train Ops Roster: use RosterExtractor (simple extraction)
            extractor = RosterExtractor(tmp_path)
            extractor.extract_tables()
            df = extractor.create_dataframe()
            sheet_name = 'Attendance Data'

        if df is None or df.empty:
            return None, "No data found in PDF. Please check the file format."

        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Clean up temporary file
        os.unlink(tmp_path)

        return output.getvalue(), None

    except Exception as e:
        return None, f"Error processing PDF: {str(e)}"


@st.cache_data(show_spinner=False)
def extract_df_from_bytes(file_bytes: bytes, department='Stations', pdf_type: str = 'attendance'):
    """Extract DataFrame from uploaded PDF bytes. Cached to make preview responsive."""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            tmp_file.write(file_bytes)
            tmp_path = tmp_file.name

        if pdf_type == 'trip_chart':
            # Trip Chart: use AttendancePDFExtractor (with Shift/Shift_Timings)
            extractor = AttendancePDFExtractor(tmp_path)
            extractor.extract_tables()
            df = extractor.create_dataframe()
        else:
            # Stations, OCC, Train Ops Roster: use RosterExtractor
            extractor = RosterExtractor(tmp_path)
            extractor.extract_tables()
            df = extractor.create_dataframe()

        try:
            os.unlink(tmp_path)
        except Exception:
            pass

        return df
    except Exception as e:
        return None


def display_roster_report(df, department, conversion_type):
    """Display roster analysis report for Train Operations Roster"""
    if department != "Train Operations" or conversion_type != "trip_chart":
        return
    
    try:
        # Generate report
        report = TrainOperationsRosterReportGenerator(df)
        
        # Display report title with date range
        if report.date_range and len(report.date_range) > 0:
            date_text = f" ({report.date_range[0]} to {report.date_range[-1]})"
        else:
            date_text = ""
        st.markdown(f"## 📊 Roster Analysis Report{date_text}")
        
        # Create tabs for different report sections
        tab1, tab2, tab3, tab4 = st.tabs(["Summary Statistics", "Daily Trends", "Shift Analysis", "Employee Details"])
        
        # Tab 1: Summary Statistics
        with tab1:
            st.markdown("### 📋 Summary Statistics")
            summary_df = report.generate_summary_table()
            st.dataframe(summary_df, use_container_width=True)
            
            # Display as metrics
            cols = st.columns(len(summary_df))
            for idx, row in summary_df.iterrows():
                with cols[idx]:
                    st.metric(row['Metric'], row['Value'])
        
        # Tab 2: Daily Trends
        with tab2:
            st.markdown("### 📈 Daily Trends")
            daily_df = report.generate_daily_trends()
            st.dataframe(daily_df, use_container_width=True)
            
            # Create visualization
            if 'Total Shifts' in daily_df.columns:
                fig = px.line(daily_df, x='Day', y='Total Shifts', 
                             markers=True, title='Daily Shift Assignments Trend')
                st.plotly_chart(fig, use_container_width=True)
        
        # Tab 3: Shift Analysis
        with tab3:
            st.markdown("### 🎯 Shift Analysis")
            
            col1, col2 = st.columns(2)
            
            shift_analysis, leave_analysis = report.generate_shift_analysis()
            
            # Shift distribution chart
            with col1:
                st.markdown("**Shift Distribution**")
                shift_df = pd.DataFrame(list(shift_analysis.items()), 
                                       columns=['Shift Type', 'Count'])
                fig1 = px.pie(shift_df, values='Count', names='Shift Type',
                             title='Working Shifts by Type')
                st.plotly_chart(fig1, use_container_width=True)
                st.dataframe(shift_df, use_container_width=True, hide_index=True)
            
            # Leave distribution chart
            with col2:
                st.markdown("**Leave Distribution**")
                leave_df = pd.DataFrame(list(leave_analysis.items()), 
                                       columns=['Leave Type', 'Count'])
                
                # Debug output
                st.write("DEBUG - Leave counts:", leave_analysis)
                
                fig2 = px.pie(leave_df, values='Count', names='Leave Type',
                             title='Leaves by Type')
                st.plotly_chart(fig2, use_container_width=True)
                st.dataframe(leave_df, use_container_width=True, hide_index=True)
        
        # Tab 4: Employee Details
        with tab4:
            st.markdown("### 👥 Employee Details Shift Analysis")
            
            employee_df = report.generate_employee_details()
            
            # Display full table
            st.dataframe(employee_df, use_container_width=True)
            
            # Summary statistics by shift type
            st.markdown("**Summary Statistics**")
            summary_cols = ['Working Shift RRTS', 'Working Shift MRTS', 'CL', 'SL', 'WL', 'CO']
            available_cols = [col for col in summary_cols if col in employee_df.columns]
            
            if available_cols:
                summary_stats = employee_df[available_cols].describe().T
                st.dataframe(summary_stats, use_container_width=True)
            
            # Top employees by shifts
            st.markdown("**Top 10 Employees by Total Assignments**")
            top_employees = employee_df.nlargest(10, 'Total')[
                ['Employee', 'Personnel_Number', 'Working Shift RRTS', 
                 'Working Shift MRTS', 'Total']
            ]
            st.dataframe(top_employees, use_container_width=True, hide_index=True)
    
    except Exception as e:
        st.error(f"Error generating report: {str(e)}")


def get_date_columns_for_month(filtered_data, year, months):
    """Generate date column headers based on year and month(s)"""
    from calendar import monthrange
    import datetime

    # Ensure year is integer
    try:
        year_int = int(year)
    except Exception:
        year_int = datetime.datetime.now().year

    # Map month names to numbers
    month_num_map = {
        'January': 1, 'February': 2, 'March': 3, 'April': 4,
        'May': 5, 'June': 6, 'July': 7, 'August': 8,
        'September': 9, 'October': 10, 'November': 11, 'December': 12,
        'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
        'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12
    }

    if not isinstance(months, list):
        months = [months]

    date_headers = []

    for month in months:
        month_upper = month.upper()
        month_num = month_num_map.get(month_upper)
        if month_num is None:
            month_num = month_num_map.get(month_upper[:3], 1)

        days_in_month = monthrange(year_int, month_num)[1]
        for day in range(1, days_in_month + 1):
            date_obj = datetime.date(year_int, month_num, day)
            date_str = date_obj.strftime("%d-%b-%Y")
            date_headers.append(date_str)

    return date_headers


def create_night_shift_report(filtered_data, year, months, department=None):
    """Create night shift report (1 for night shift, 0 otherwise)

    Falls back to minimal behavior if employee details loader is not available.
    """
    # Try to load employee details helper if available
    emp_details = None
    try:
        from analyzer import load_employee_details
        emp_details = load_employee_details(department)
    except Exception:
        emp_details = None

    # Ensure months is list
    if not isinstance(months, list):
        months = [months]

    # Identify unique employees
    if 'Personnel_Number' in filtered_data.columns:
        unique_employees = filtered_data.drop_duplicates(subset=['Personnel_Number'])
    else:
        unique_employees = filtered_data.drop_duplicates()

    day_columns = [col for col in filtered_data.columns if col.startswith('Day_')]
    date_headers = get_date_columns_for_month(filtered_data, year, months)

    # Map month to its subset
    month_data_map = {}
    if 'Month' in filtered_data.columns:
        for month in months:
            month_data_map[month] = filtered_data[filtered_data['Month'] == month]
    else:
        month_data_map[months[0]] = filtered_data

    report_data = []

    for idx, emp_row in unique_employees.iterrows():
        raw_emp_id = str(emp_row.get('Personnel_Number', ''))
        # Clean the ID (remove .0 and whitespaces) to ensure robust matching
        emp_id = raw_emp_id.split('.')[0].strip()
        
        if department == "Train Operations":
            loc_val = emp_details['to_ta'].get(emp_id, '') if emp_details and 'to_ta' in emp_details else ''
        else:
            loc_val = emp_details['station'].get(emp_id, '') if emp_details and 'station' in emp_details else ''
            
        emp_record = {
            'Employee': emp_row.get('Employee', ''),
            'Personnel_Number': emp_id,
            'Designation': emp_details['designation'].get(emp_id, 'N/A') if emp_details and 'designation' in emp_details else 'N/A',
        }
        
        if department == "Train Operations":
            emp_record['Crew Control'] = loc_val
        else:
            emp_record['Station'] = loc_val
            
        emp_record['AM'] = emp_details['am'].get(emp_id, '') if emp_details and 'am' in emp_details else ''

        date_idx = 0
        for month in months:
            month_data = month_data_map.get(month, pd.DataFrame())

            if month_data.empty:
                from calendar import monthrange
                month_num_map = {
                    'January': 1, 'February': 2, 'March': 3, 'April': 4,
                    'May': 5, 'June': 6, 'July': 7, 'August': 8,
                    'September': 9, 'October': 10, 'November': 11, 'December': 12,
                    'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
                    'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12
                }
                month_upper = month.upper()
                month_num = month_num_map.get(month_upper)
                if month_num is None:
                    month_num = month_num_map.get(month_upper[:3], 1)
                days_in_month = monthrange(int(year), month_num)[1]

                for _ in range(days_in_month):
                    if date_idx < len(date_headers):
                        emp_record[date_headers[date_idx]] = 0
                        date_idx += 1
            else:
                emp_month_data = month_data[month_data.get('Personnel_Number', pd.Series()).astype(str) == str(emp_row.get('Personnel_Number', ''))]
                if emp_month_data.empty:
                    from calendar import monthrange
                    month_num_map = {
                        'January': 1, 'February': 2, 'March': 3, 'April': 4,
                        'May': 5, 'June': 6, 'July': 7, 'August': 8,
                        'September': 9, 'October': 10, 'November': 11, 'December': 12,
                        'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
                        'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12
                    }
                    month_upper = month.upper()
                    month_num = month_num_map.get(month_upper)
                    if month_num is None:
                        month_num = month_num_map.get(month_upper[:3], 1)
                    days_in_month = monthrange(int(year), month_num)[1]

                    for _ in range(days_in_month):
                        if date_idx < len(date_headers):
                            emp_record[date_headers[date_idx]] = 0
                            date_idx += 1
                else:
                    emp_month_row = emp_month_data.iloc[0]
                    for day_col in day_columns:
                        if date_idx >= len(date_headers):
                            break
                        date_header = date_headers[date_idx]
                        cell_value = emp_month_row.get(day_col, '')
                        if pd.notna(cell_value) and str(cell_value).strip():
                            cell_str = str(cell_value).strip().upper()
                            if cell_str.startswith('N') or cell_str.startswith('N-'):
                                emp_record[date_header] = 1
                            else:
                                emp_record[date_header] = 0
                        else:
                            emp_record[date_header] = 0
                        date_idx += 1

        report_data.append(emp_record)

    report_df = pd.DataFrame(report_data)
    info_cols = ['Employee', 'Personnel_Number', 'Designation', 'Station', 'Crew Control', 'AM']
    date_cols = [col for col in report_df.columns if col not in info_cols]
    report_df['Total'] = report_df[date_cols].sum(axis=1) if date_cols else 0

    # Sort by AM and Employee
    sort_cols = [c for c in ['AM', 'Employee'] if c in report_df.columns]
    if sort_cols:
        report_df = report_df.sort_values(sort_cols)

    return report_df


def main():
    """Main application function"""

    # Header
    st.markdown('<div class="main-header">📊 Attendance Analysis Tool</div>', unsafe_allow_html=True)

    # Sidebar (same structure as dashboard)
    st.sidebar.title("⚙️ Tool Controls")

    # Department Selection (for future use)
    st.sidebar.header("1️⃣ Department")
    department = st.sidebar.selectbox(
        "Choose Department",
        ["Stations", "OCC", "Train Operations"],
        help="Select the department (for future features)"
    )

    # Train Operations specific option
    train_ops_mode = None
    if department == "Train Operations":
        st.sidebar.header("1b️⃣ Train Operations Type")
        train_ops_mode = st.sidebar.radio(
            "Select conversion type",
            ["Trip Chart", "Roster"],
            help="Choose what type of PDF to convert"
        )

    # Time Period
    st.sidebar.header("2️⃣ Time Period")
    selected_year = st.sidebar.selectbox("Year", [2024, 2025, 2026], index=1)
    available_months = ["January", "February", "March", "April", "May", "June",
                       "July", "August", "September", "October", "November", "December"]
    selected_months = st.sidebar.multiselect(
        "Months",
        available_months,
        default=[],
        help="Select months (required for Night Shift Report)"
    )

    # Analysis Options
    st.sidebar.header("3️⃣ Analysis Options")
    show_summary = st.sidebar.checkbox("📊 Summary Statistics", value=False, disabled=True)
    show_daily = st.sidebar.checkbox("📅 Daily Trends", value=False, disabled=True)
    show_performance = st.sidebar.checkbox("🎯 Performance Analysis", value=False, disabled=True)
    show_shifts = st.sidebar.checkbox("🔄 Shift Analysis", value=False, disabled=True)
    show_employees = st.sidebar.checkbox("👥 Employee Details", value=False, disabled=True)
    show_night_shift_report = st.sidebar.checkbox("🌙 Night Shift Report", value=False, help="View night shift report with 1/0 notation and date columns")

    # Alert Thresholds (for future use)
    st.sidebar.header("4️⃣ Alert Thresholds")
    threshold_80 = st.sidebar.slider("Warning Threshold (%)", 0, 100, 80, disabled=True)
    threshold_90 = st.sidebar.slider("Good Threshold (%)", 0, 100, 90, disabled=True)

    # Data Management (for future use)
    st.sidebar.header("5️⃣ Data Management")
    st.sidebar.info("🔄 Future: Refresh from PDFs")

    # Main Content - PDF to Excel Converter
    st.markdown("## 📄 PDF to Excel Converter")
    
    # Determine conversion type based on department and mode
    if department == "Train Operations" and train_ops_mode == "Trip Chart":
        st.markdown("Upload your Trip Chart PDF file and download the converted Excel file.")
        file_type_hint = "Trip Chart PDF"
        conversion_type = "attendance"
    elif department == "Train Operations" and train_ops_mode == "Roster":
        st.markdown("Upload your Train Operations Roster PDF file and download the converted Excel file.")
        file_type_hint = "Roster PDF"
        conversion_type = "trip_chart"  # Use simple roster extraction
    else:
        st.markdown("Upload your attendance PDF file and download the converted Excel file.")
        file_type_hint = "Attendance PDF"
        conversion_type = "attendance"  # Use simple roster extraction

    # File upload
    uploaded_file = st.file_uploader(
        "Choose a PDF file",
        type=['pdf'],
        help=f"Upload a {file_type_hint} file"
    )

    if uploaded_file is not None:
        st.success(f"✅ File uploaded: {uploaded_file.name}")

        # Preview (optional) - allow preview before or after conversion
        if st.checkbox("👀 Preview Data"):
            with st.spinner("Extracting preview..."):
                df = extract_df_from_bytes(uploaded_file.getvalue(), department, conversion_type)

            if df is not None and not df.empty:
                st.markdown("### 📋 Data Preview")
                st.dataframe(df.head(20), use_container_width=True)
                st.info(f"📊 Total rows: {len(df)} | Total columns: {len(df.columns)}")

            else:
                st.warning("No data extracted for preview. Please check the PDF format.")

        # Process button
        if st.button("🔄 Convert to Excel", type="primary"):
            with st.spinner("Processing PDF... This may take a few moments."):
                excel_data, error = process_pdf_to_excel(uploaded_file, department, conversion_type)

            if error:
                st.error(f"❌ {error}")
            else:
                st.success("✅ PDF converted successfully!")

                # Download button
                if department == "Train Operations" and train_ops_mode == "Trip Chart":
                    file_name = f"{Path(uploaded_file.name).stem}_trip_chart.xlsx"
                elif department == "Train Operations" and train_ops_mode == "Roster":
                    file_name = f"{Path(uploaded_file.name).stem}_roster.xlsx"
                else:
                    file_name = f"{Path(uploaded_file.name).stem}_converted.xlsx"

                st.download_button(
                    label="📊 Download Excel File",
                    data=excel_data,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download the converted Excel file"
                )
                
                # Generate and display report for Train Operations Roster
                if department == "Train Operations" and train_ops_mode == "Roster":
                    st.markdown("---")
                    df = extract_df_from_bytes(uploaded_file.getvalue(), department, conversion_type)
                    if df is not None and not df.empty:
                        display_roster_report(df, department, conversion_type)

                # Generate and display Night Shift Report
                if show_night_shift_report:
                    if not selected_months:
                        st.warning("⚠️ Please select at least one month from the sidebar 'Time Period' for the Night Shift Report.")
                    else:
                        st.markdown("---")
                        st.markdown("### 🌙 Night Shift Report")
                        df = extract_df_from_bytes(uploaded_file.getvalue(), department, conversion_type)
                        
                        if df is not None and not df.empty:
                            with st.spinner("Generating Night Shift Report..."):
                                night_report_df = create_night_shift_report(df, selected_year, selected_months, department)
                                
                                info_cols = ['Employee', 'Personnel_Number', 'Designation', 'Station', 'Crew Control', 'AM', 'Total']
                                date_cols = [col for col in night_report_df.columns if col not in info_cols]
                                
                                st.info(f"📅 Showing {len(date_cols)} days | 🌙 Red cells = Night Shift (value 1) | Total column shows count of night shifts")
                                
                                def highlight_night_shifts(val):
                                    if val == 1:
                                        return 'background-color: #ffcccc; font-weight: bold; color: black;'
                                    return ''
                                
                                try:
                                    styled_report = night_report_df.style.map(highlight_night_shifts, subset=date_cols)
                                except AttributeError:
                                    styled_report = night_report_df.style.applymap(highlight_night_shifts, subset=date_cols)
                                
                                st.dataframe(styled_report, use_container_width=True, height=600)
                                
                                # Export options
                                col_exp1, col_exp2 = st.columns(2)
                                month_str_file = "_".join(selected_months)
                                
                                with col_exp1:
                                    csv_data = night_report_df.to_csv(index=False).encode('utf-8')
                                    st.download_button(
                                        label="⬇️ Download Night Shift Report as CSV",
                                        data=csv_data,
                                        file_name=f"{Path(uploaded_file.name).stem}_night_shift_{selected_year}_{month_str_file}.csv",
                                        mime="text/csv",
                                        key=f"csv_{month_str_file}_night_shift"
                                    )
                                
                                with col_exp2:
                                    output = io.BytesIO()
                                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                        night_report_df.to_excel(writer, sheet_name='Night Shift Report', index=False)
                                        try:
                                            from openpyxl.styles import PatternFill
                                            worksheet = writer.sheets['Night Shift Report']
                                            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                                            header_row = list(night_report_df.columns)
                                            if date_cols:
                                                first_date_col_idx = header_row.index(date_cols[0]) + 1
                                                last_date_col_idx = header_row.index(date_cols[-1]) + 1
                                                from openpyxl.utils import get_column_letter
                                                first_col_letter = get_column_letter(first_date_col_idx)
                                                last_col_letter = get_column_letter(last_date_col_idx)
                                                
                                                for row in worksheet.iter_rows(min_row=2, min_col=first_date_col_idx, max_col=last_date_col_idx):
                                                    for cell in row:
                                                        if cell.value == 1:
                                                            cell.fill = red_fill
                                        except Exception as e:
                                            pass
                                            
                                    st.download_button(
                                        label="📊 Download Night Shift Report as Excel",
                                        data=output.getvalue(),
                                        file_name=f"{Path(uploaded_file.name).stem}_night_shift_{selected_year}_{month_str_file}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"excel_{month_str_file}_night_shift"
                                    )

    # Instructions
    with st.expander("📖 How to Use"):
        st.markdown("""
        1. **Upload PDF**: Select your attendance report PDF file using the file uploader above
        2. **Convert**: Click the "Convert to Excel" button to process the PDF
        3. **Download**: Once conversion is complete, download the Excel file
        4. **Preview**: Optionally check the data preview to verify the conversion

        **Supported Formats**: PDF files containing attendance tables with employee data
        """)

    # Footer
    st.markdown("---")
    st.markdown(
        f"**Last Updated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
        "**Version:** Online Tool v1.0"
    )


if __name__ == "__main__":
    main()